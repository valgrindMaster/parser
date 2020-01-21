from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import column_index_from_string

# Load spreadsheet for data processing.
wb = load_workbook('Data1.xlsx')

# Get spreadsheet.
ws = wb.worksheets[0]

# Mark all header starting indexes.
headerStartingIndexes = []

for cell in ws[1]:
  if cell.value:
    headerStartingIndexes.append(cell.column)

# There are certain headers that actually utilize their full width of columns, so skip those.
headerExceptions = [column_index_from_string('AL')]
columnOffset = 0

headerStartingIndexesLen = len(headerStartingIndexes)

for i in range(headerStartingIndexesLen):
  print(str(round((i / headerStartingIndexesLen) * 100, 2)), '% Complete.', end='\r')

  colStart = headerStartingIndexes[i] - columnOffset
  colEnd = ((ws.max_column if (i == headerStartingIndexesLen - 1) else headerStartingIndexes[i + 1]) - 1) - columnOffset

  if abs(colStart - colEnd) == 0 or colStart + columnOffset in headerExceptions:
    continue

  for j in range(2, ws.max_row):
    combinedCellValues = set()
    for k in range(colStart, colEnd):
      value = ws.cell(row = j, column = k + 1).value
      if value != None:
        value = str(value).strip().replace(r'^[\r\n\t]*\b|\b[\r\n\t]*$', '')
        combinedCellValues.add(value)
    if len(combinedCellValues) > 0:
      ws.cell(row = j, column = colStart, value = (', '.join(combinedCellValues)))
  
  delColStart = colStart + 1
  amount = abs(colStart - colEnd)
  ws.delete_cols(delColStart, amount)
  columnOffset += amount

wb.save('final.xlsx')